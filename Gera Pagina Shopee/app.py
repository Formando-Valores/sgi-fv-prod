import os
import io
import uuid
from datetime import datetime
from typing import List, Dict, Any

from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from werkzeug.utils import secure_filename

# -----------------------
# google ai 
# -----------------------
from google import genai
from google.genai import types

import boto3
from PIL import Image, ImageFilter, ImageEnhance
from openpyxl import load_workbook

# -----------------------
# Config
# -----------------------
SD_API_URL = os.getenv("SD_API_URL", "http://127.0.0.1:7860")
SD_MODEL = os.getenv("SD_MODEL", "")  # opcional: se quiser trocar checkpoint via API depois

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret")  # só pra flash messages local

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
GEN_DIR = os.path.join(BASE_DIR, "generated")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(GEN_DIR, exist_ok=True)

TEMPLATE_XLSX = os.path.join(BASE_DIR, "Shopee_mass_upload_2026-02-25_basic_template.xlsx")
SHEET_NAME = "Modelo"

S3_ENDPOINT = os.getenv("SUPABASE_S3_ENDPOINT")
S3_REGION = os.getenv("SUPABASE_REGION", "us-west-2")
S3_ACCESS_KEY = os.getenv("SUPABASE_ACCESS_KEY")
S3_SECRET_KEY = os.getenv("SUPABASE_SECRET_KEY")
S3_BUCKET = os.getenv("SUPABASE_BUCKET", "Produtos")

MODEL_EXPORT_XLSX = os.path.join(BASE_DIR, "Modelo a ser gerado.xlsx")
MODEL_SHEET_NAME = "Planilha1"  # do seu arquivo

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GEMINI_IMAGE_MODEL = os.getenv("GEMINI_IMAGE_MODEL", "gemini-2.5-flash-image")  # ou preview

PUBLIC_BASE = "https://axxjqbunfpnhimdhhxms.storage.supabase.co/storage/v1/object/public"

# “Banco” simples em memória (pra uso local).
# Se você reiniciar o servidor, limpa.
PRODUCTS: List[Dict[str, Any]] = []


import base64
import requests
import io
from PIL import Image

def sd_generate_background(prompt: str, width=1024, height=1024, steps=24, cfg=6.5, seed=-1) -> Image.Image:
    payload = {
        "prompt": prompt,
        "negative_prompt": "blurry, low quality, distorted, watermark, text, logo, extra objects, messy",
        "steps": steps,
        "cfg_scale": cfg,
        "width": width,
        "height": height,
        "seed": seed,
        "sampler_name": "Euler a",
    }

    r = requests.post(f"{SD_API_URL}/sdapi/v1/txt2img", json=payload, timeout=300)
    r.raise_for_status()
    data = r.json()

    # a API retorna imagens em base64
    b64 = data["images"][0]
    img_bytes = base64.b64decode(b64.split(",", 1)[-1])
    return Image.open(io.BytesIO(img_bytes)).convert("RGB")


from PIL import ImageFilter

def compose_product_on_background(product_png_path: str, background_rgb: Image.Image) -> Image.Image:
    prod = Image.open(product_png_path).convert("RGBA")
    bg = background_rgb.convert("RGBA")

    bw, bh = bg.size
    pw, ph = prod.size

    # produto ocupando ~70% do canvas
    scale = min((bw * 0.70) / pw, (bh * 0.70) / ph)
    new_size = (max(1, int(pw * scale)), max(1, int(ph * scale)))
    prod_r = prod.resize(new_size, Image.LANCZOS)

    x = (bw - new_size[0]) // 2
    y = (bh - new_size[1]) // 2 + int(bh * 0.06)

    # sombra pelo alpha
    alpha = prod_r.split()[-1]
    shadow = Image.new("RGBA", prod_r.size, (0, 0, 0, 180))
    shadow.putalpha(alpha)
    shadow = shadow.filter(ImageFilter.GaussianBlur(18))

    out = bg.copy()
    out.alpha_composite(shadow, (x, y + 28))
    out.alpha_composite(prod_r, (x, y))
    return out.convert("RGB")
    
def sd_generate_lifestyle_images(product_png_path: str, categoria_nome: str) -> list[Image.Image]:
    scenes = CATEGORY_SCENES.get(categoria_nome) or DEFAULT_SCENES
    outs = []
    for scene_prompt in scenes[:3]:
        bg = sd_generate_background(scene_prompt, width=1024, height=1024, steps=24, cfg=6.5)
        final_img = compose_product_on_background(product_png_path, bg)
        outs.append(final_img)
    return outs


# -----------------------
# Supabase S3 client
# -----------------------
def s3_client():
    if not (S3_ENDPOINT and S3_ACCESS_KEY and S3_SECRET_KEY):
        raise RuntimeError("Defina SUPABASE_S3_ENDPOINT / SUPABASE_ACCESS_KEY / SUPABASE_SECRET_KEY nas variáveis de ambiente.")
    return boto3.client(
        "s3",
        endpoint_url=S3_ENDPOINT,
        aws_access_key_id=S3_ACCESS_KEY,
        aws_secret_access_key=S3_SECRET_KEY,
        region_name=S3_REGION,
    )


def upload_jpeg_and_get_public_url(client, bucket: str, key: str, img_rgb: Image.Image) -> str:
    buf = io.BytesIO()
    img_rgb.save(buf, format="JPEG", quality=92, optimize=True)
    buf.seek(0)

    client.put_object(
        Bucket=bucket,
        Key=key,
        Body=buf.getvalue(),
        ContentType="image/jpeg",
        ACL="public-read",
    )
    return f"{PUBLIC_BASE}/{bucket}/{key}"


# -----------------------
# Imagens (modo C)
# -----------------------
def remove_background_rgba(img: Image.Image) -> Image.Image:
    from rembg import remove
    out = remove(img)
    if isinstance(out, (bytes, bytearray)):
        out = Image.open(io.BytesIO(out))
    return out.convert("RGBA")


def fit_product_on_canvas(product_rgba: Image.Image, canvas_size=(2000, 2000), margin=180) -> Image.Image:
    w, h = product_rgba.size
    cw, ch = canvas_size
    max_w = cw - 2 * margin
    max_h = ch - 2 * margin
    scale = min(max_w / w, max_h / h)
    new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
    product_resized = product_rgba.resize(new_size, Image.LANCZOS)

    canvas = Image.new("RGBA", canvas_size, (0, 0, 0, 0))
    x = (cw - new_size[0]) // 2
    y = (ch - new_size[1]) // 2
    canvas.alpha_composite(product_resized, (x, y))
    return canvas


def make_soft_shadow(product_canvas_rgba: Image.Image, blur=18, opacity=0.35, y_offset=60) -> Image.Image:
    alpha = product_canvas_rgba.split()[-1]
    shadow = Image.new("RGBA", product_canvas_rgba.size, (0, 0, 0, 0))
    shadow_mask = alpha.point(lambda a: int(a * opacity))
    shadow.putalpha(shadow_mask)
    shadow = shadow.filter(ImageFilter.GaussianBlur(blur))

    shifted = Image.new("RGBA", product_canvas_rgba.size, (0, 0, 0, 0))
    shifted.alpha_composite(shadow, (0, y_offset))
    return shifted


def bg_white(size):
    return Image.new("RGBA", size, (255, 255, 255, 255))


def bg_gradient_soft(size):
    w, h = size
    base = Image.new("RGBA", size, (255, 255, 255, 255))
    grad = Image.new("L", (1, h))
    for y in range(h):
        v = int(255 - (20 * (y / max(1, h - 1))))
        grad.putpixel((0, y), v)
    grad = grad.resize((w, h))
    overlay = Image.merge("RGBA", (grad, grad, grad, Image.new("L", (w, h), 255)))
    return Image.alpha_composite(base, overlay)


def bg_lifestyle_clean(size):
    w, h = size
    wall = Image.new("RGBA", (w, h), (245, 245, 245, 255))
    table_h = int(h * 0.28)
    table = Image.new("RGBA", (w, table_h), (235, 235, 235, 255))
    wall.alpha_composite(table, (0, h - table_h))
    line = Image.new("RGBA", (w, 3), (220, 220, 220, 255))
    wall.alpha_composite(line, (0, h - table_h))
    return wall


def compose_variant(product_canvas_rgba: Image.Image, bg: Image.Image, add_shadow=True) -> Image.Image:
    out = bg.copy()
    if add_shadow:
        out = Image.alpha_composite(out, make_soft_shadow(product_canvas_rgba))
    out = Image.alpha_composite(out, product_canvas_rgba)

    out = ImageEnhance.Contrast(out).enhance(1.03)
    out = ImageEnhance.Sharpness(out).enhance(1.05)
    return out.convert("RGB")


def generate_3_variants_from_file(image_path: str) -> List[Image.Image]:
    original = Image.open(image_path).convert("RGBA")
    cut = remove_background_rgba(original)
    canvas = fit_product_on_canvas(cut, canvas_size=(2000, 2000), margin=180)

    imgs = [
        compose_variant(canvas, bg_white(canvas.size), add_shadow=True),
        compose_variant(canvas, bg_gradient_soft(canvas.size), add_shadow=True),
        compose_variant(canvas, bg_lifestyle_clean(canvas.size), add_shadow=True),
    ]
    return imgs


# -----------------------
# Excel (template Shopee)
# -----------------------
def build_header_map(ws, header_row=3) -> Dict[str, int]:
    header_map = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v:
            header_map[str(v).strip()] = col
    return header_map


def set_cell(ws, header_map: Dict[str, int], row: int, header: str, value):
    col = header_map.get(header)
    if col is None:
        return
    ws.cell(row=row, column=col).value = value


def fill_product_row(ws, header_map: Dict[str, int], row: int, produto: Dict[str, Any]):
    set_cell(ws, header_map, row, "Categoria", str(produto["categoria_id"]))
    set_cell(ws, header_map, row, "Nome do Produto", produto["nome"])
    set_cell(ws, header_map, row, "Descrição do Produto", produto["descricao"])
    set_cell(ws, header_map, row, "SKU principal", produto["sku"])
    set_cell(ws, header_map, row, "Preço", float(produto["preco"]))
    set_cell(ws, header_map, row, "Estoque", int(produto["estoque"]))

    # capa + 1/2/3
    urls = produto["urls"]
    if urls:
        set_cell(ws, header_map, row, "Imagem de capa", urls[0])
    if len(urls) > 1:
        set_cell(ws, header_map, row, "Imagem do produto 1", urls[1])
    if len(urls) > 2:
        set_cell(ws, header_map, row, "Imagem do produto 2", urls[2])
    if len(urls) > 3:
        set_cell(ws, header_map, row, "Imagem do produto 3", urls[3])
        
# Criar cliente Google Ai
def gemini_client():
    if not GEMINI_API_KEY:
        raise RuntimeError("Defina GEMINI_API_KEY nas variáveis de ambiente.")
    return genai.Client(api_key=GEMINI_API_KEY)
    
CATEGORY_SCENES = {
    "Decoração": [
        "photorealistic modern minimalist living room, neutral tones, soft daylight, shallow depth of field, no text, no watermark",
        "photorealistic cozy dining table scene, warm lighting, neutral decor, no text, no watermark",
        "photorealistic scandinavian shelf/sideboard interior, bright airy, no text, no watermark",
    ],
    "Cozinha": [
        "photorealistic modern kitchen countertop, marble surface, daylight, clean scene, no text, no watermark",
        "photorealistic table setting (mesa posta), soft natural light, clean style, no text, no watermark",
        "photorealistic bright rustic kitchen background, light wood, neutral colors, no text, no watermark",
    ],
    "Eletrônicos": [
        "photorealistic modern home office desk setup, soft daylight, clean, no text, no watermark",
        "photorealistic minimalist workstation, neutral background, shallow depth of field, no text, no watermark",
        "photorealistic cozy desk setup, warm light, modern, no text, no watermark",
    ],
}

DEFAULT_SCENES = [
    "photorealistic clean studio background, soft gradient light, no text, no watermark",
    "photorealistic modern neutral interior background, soft daylight, no text, no watermark",
    "photorealistic minimalist scene, neutral colors, no text, no watermark",
]



def gemini_generate_lifestyle_images(product_png_path: str, categoria: str) -> List[Image.Image]:
    """
    Gera exatamente 3 imagens.
    A imagem enviada é APENAS referência (não entra como '4ª imagem').
    """
    client = gemini_client()

    # abre a imagem referência (de preferência PNG recortado)
    product_img = Image.open(product_png_path).convert("RGBA")

    scenes = CATEGORY_SCENES.get(categoria) or [
        "a clean white studio background, professional product photography",
        "a minimal lifestyle indoor scene, soft daylight, product naturally placed",
        "a modern neutral interior scene, realistic lighting and shadows, product featured",
    ]

    base_instruction = (
        "Use the provided product image as the exact product reference. "
        "DO NOT change the product: no changes to shape, color, logo, text, proportions, or details. "
        "Only change the background / environment to match the scene. "
        "Place the product realistically in the scene and add realistic natural shadow. "
        "Photorealistic product photography, high quality."
    )

    results: List[Image.Image] = []

    for scene in scenes[:3]:
        prompt = f"{base_instruction}\nScene: {scene}"

        # gera/edita imagem (prompt + imagem)
        response = client.models.generate_content(
            model=GEMINI_IMAGE_MODEL,
            contents=[prompt, product_img],
        )

        # pega a imagem gerada do retorno (parts inline_data)
        out_img = None
        for part in response.candidates[0].content.parts:
            if getattr(part, "inline_data", None) is not None:
                out_img = Image.open(io.BytesIO(part.inline_data.data)).convert("RGB")
                break

        if out_img is None:
            raise RuntimeError("Gemini não retornou imagem. Verifique modelo/permissões.")

        results.append(out_img)

    return results


from openpyxl import Workbook

from openpyxl import Workbook
from datetime import datetime
import os

from openpyxl import Workbook, load_workbook

def export_shopee_xlsx(products: List[Dict[str, Any]]) -> str:
    # 1) Ler o modelo (seu arquivo limpo)
    model_wb = load_workbook(MODEL_EXPORT_XLSX, data_only=True)
    model_ws = model_wb[MODEL_SHEET_NAME]

    # Cabeçalhos (linha 1) e obrigatoriedade (linha 2)
    headers = []
    required_flags = []
    max_col = model_ws.max_column

    for col in range(1, max_col + 1):
        headers.append(model_ws.cell(row=1, column=col).value)
        required_flags.append(model_ws.cell(row=2, column=col).value)

    # 2) Definir quais colunas exportar:
    # - tudo que for "Obrigatório"
    # - + campos essenciais para publicação/importação
    essential_headers = {
        "Categoria",
        "SKU principal",
        "Estoque",
        "Imagem de capa",
        "Imagem do produto 1",
        "Imagem do produto 2",
        "Imagem do produto 3",
        "Comprimento",
        "Largura",
        "Altura",
        "Correios",
        "Prazo de Postagem para Encomenda",
    }

    selected_cols = []
    for idx, h in enumerate(headers, start=1):
        if not h:
            continue
        flag = required_flags[idx - 1]
        if flag == "Obrigatório" or h in essential_headers:
            selected_cols.append(idx)  # índice da coluna no modelo

    # 3) Criar workbook novo e copiar linhas 1–4 do modelo (somente colunas selecionadas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    for target_col, source_col in enumerate(selected_cols, start=1):
        for r in range(1, 5):  # linhas 1..4
            ws.cell(row=r, column=target_col).value = model_ws.cell(row=r, column=source_col).value

    # Criar mapa header->coluna no arquivo novo (linha 1)
    header_map = {}
    for c in range(1, len(selected_cols) + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            header_map[str(v).strip()] = c

    # Helper para setar célula por header
    def setv(row: int, header: str, value):
        col = header_map.get(header)
        if col:
            ws.cell(row=row, column=col).value = value

    # 4) Escrever produtos a partir da linha 5
    start_row = 5
    for i, p in enumerate(products):
        row = start_row + i

        setv(row, "Categoria", p.get("categoria_id", ""))
        setv(row, "Nome do Produto", p.get("nome", ""))
        setv(row, "Descrição do Produto", p.get("descricao", ""))
        setv(row, "SKU principal", p.get("sku", ""))

        setv(row, "Preço", p.get("preco", 0))
        setv(row, "Estoque", p.get("estoque", 0))

        # peso e dimensões
        setv(row, "Peso", p.get("peso", ""))  # obrigatório no seu modelo
        setv(row, "Comprimento", p.get("comprimento", ""))
        setv(row, "Largura", p.get("largura", ""))
        setv(row, "Altura", p.get("altura", ""))

        # logística
        setv(row, "Correios", p.get("correios", ""))
        setv(row, "Prazo de Postagem para Encomenda", p.get("prazo_postagem", ""))

        # imagens (URLs)
        urls = p.get("urls", [])
        if urls:
            setv(row, "Imagem de capa", urls[0])
        if len(urls) > 1:
            setv(row, "Imagem do produto 1", urls[1])
        if len(urls) > 2:
            setv(row, "Imagem do produto 2", urls[2])
        if len(urls) > 3:
            setv(row, "Imagem do produto 3", urls[3])

    # 5) Salvar
    filename = f"shopee_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(GEN_DIR, filename)
    wb.save(out_path)
    return out_path


# -----------------------
# Rotas Web
# -----------------------
@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", products=PRODUCTS)


@app.route("/add", methods=["POST"])
@app.route("/add", methods=["POST"])
def add_product():
    try:
        nome = request.form.get("nome", "").strip()
        categoria_nome = request.form.get("categoria_nome", "").strip()  # usado pra cenários Gemini
        categoria_id = request.form.get("categoria_id", "").strip()      # opcional: pra colocar na planilha, se existir
        descricao = request.form.get("descricao", "").strip()

        preco_raw = request.form.get("preco", "").strip()
        estoque_raw = request.form.get("estoque", "0").strip()

        peso_raw = request.form.get("peso", "").strip()
        comprimento_raw = request.form.get("comprimento", "").strip()
        largura_raw = request.form.get("largura", "").strip()
        altura_raw = request.form.get("altura", "").strip()

        correios = request.form.get("correios", "").strip()
        prazo_postagem = request.form.get("prazo_postagem", "").strip()

        file = request.files.get("imagem")

        # validações mínimas (ajuste como quiser)
        if not nome or not categoria_nome or not descricao or not file:
            flash("Preencha nome, categoria (nome), descrição e envie uma imagem.")
            return redirect(url_for("index"))

        if not preco_raw:
            flash("Preço é obrigatório.")
            return redirect(url_for("index"))

        if not peso_raw:
            flash("Peso é obrigatório.")
            return redirect(url_for("index"))

        # conversões numéricas (aceita vírgula ou ponto)
        def to_float(s: str) -> float:
            return float(s.replace(",", ".").strip())

        def to_int(s: str) -> int:
            s = s.strip()
            return int(s) if s else 0

        preco = to_float(preco_raw)
        peso = to_float(peso_raw)

        estoque = to_int(estoque_raw)

        comprimento = to_float(comprimento_raw) if comprimento_raw else ""
        largura = to_float(largura_raw) if largura_raw else ""
        altura = to_float(altura_raw) if altura_raw else ""

        # salvar upload
        safe_name = secure_filename(file.filename)
        local_id = uuid.uuid4().hex[:10]
        local_path = os.path.join(UPLOAD_DIR, f"{local_id}_{safe_name}")
        file.save(local_path)

        # recorte (referência pro Gemini)
        original = Image.open(local_path).convert("RGBA")
        cut = remove_background_rgba(original)

        cut_path = os.path.join(UPLOAD_DIR, f"{local_id}_cut.png")
        cut.save(cut_path, "PNG")

        # gerar 3 imagens (Gemini) baseado na categoria_nome
        imgs = sd_generate_lifestyle_images(cut_path, categoria_nome)

        if not imgs or len(imgs) != 3:
            raise RuntimeError("Gemini não retornou exatamente 3 imagens.")

        # upload pro Supabase
        client = s3_client()
        slug = nome.lower().replace(" ", "_")[:40]
        unique = uuid.uuid4().hex[:8]
        folder = f"catalogo/{slug}-{unique}"

        keys = [
            f"{folder}/cover_1.jpg",
            f"{folder}/img_2.jpg",
            f"{folder}/img_3.jpg",
        ]

        urls = [
            upload_jpeg_and_get_public_url(client, S3_BUCKET, keys[0], imgs[0]),
            upload_jpeg_and_get_public_url(client, S3_BUCKET, keys[1], imgs[1]),
            upload_jpeg_and_get_public_url(client, S3_BUCKET, keys[2], imgs[2]),
        ]

        produto = {
            "nome": nome,
            "categoria_nome": categoria_nome,    # pra logs/lista
            "categoria_id": categoria_id,        # se você quiser exportar também
            "descricao": descricao,
            "preco": preco,
            "estoque": estoque,
            "peso": peso,
            "comprimento": comprimento,
            "largura": largura,
            "altura": altura,
            "correios": correios,
            "prazo_postagem": prazo_postagem,
            "sku": f"SKU-{uuid.uuid4().hex[:10].upper()}",
            "upload_path": local_path,
            "urls": urls,
        }

        PRODUCTS.append(produto)
        flash(f"Produto adicionado: {nome}")
        return redirect(url_for("index"))

    except Exception as e:
        flash(f"Erro ao adicionar produto: {e}")
        return redirect(url_for("index"))


@app.route("/remove/<int:index>", methods=["POST"])
def remove_product(index: int):
    if 0 <= index < len(PRODUCTS):
        removed = PRODUCTS.pop(index)
        flash(f"Removido: {removed.get('nome')}")
    return redirect(url_for("index"))


@app.route("/export", methods=["POST"])
def export():
    if not PRODUCTS:
        flash("Nenhum produto na lista para exportar.")
        return redirect(url_for("index"))

    try:
        out_path = export_shopee_xlsx(PRODUCTS)
        return send_file(out_path, as_attachment=True, download_name=os.path.basename(out_path))
    except Exception as e:
        flash(f"Erro ao exportar planilha: {e}")
        return redirect(url_for("index"))


@app.route("/clear", methods=["POST"])
def clear():
    PRODUCTS.clear()
    flash("Lista de produtos limpa.")
    return redirect(url_for("index"))


if __name__ == "__main__":
    # servidor local
    app.run(host="127.0.0.1", port=5000, debug=True)